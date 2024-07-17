import docx 
import openpyxl
import argparse

parser = argparse.ArgumentParser(description="Modify the data in a xlsx file from a docx file")
parser.add_argument("-o", "--output", default= "./modified_data.xlsx", help="Output xlsx file path")
parser.add_argument("-s", "--source", default= "./data.xlsx", help="Source xlsx file path")
parser.add_argument("-i", "--input", default= "./data.docx", help="Input doc file path")
args = parser.parse_args()
docc = docx.Document(args.input)
namelist = []
for data in docc.paragraphs:
    namelist.append((data.text.rsplit(' ')[0], data.text.rsplit(' ')[1]))

wb = openpyxl.open(args.source)
ws = wb.active
for adjust_object in namelist:
    for row in range(1, 4):
            cell = ws.cell(row, 1)
            
            if cell.value == adjust_object[0]:
                print(f"Modified Data:  {ws.cell(row, column=1).value}   {ws.cell(row, column=2).value} -> {adjust_object[1]}")
                ws.cell(row, column=2).value = adjust_object[1]
                

wb.save(args.output)